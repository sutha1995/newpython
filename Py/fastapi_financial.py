from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
import io
from fastapi import FastAPI, HTTPException, status
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel, EmailStr
from typing import List, Optional, Dict
from datetime import datetime
import csv
import sqlite3
from pathlib import Path
from financial_statement_sqlite import FinancialDatabaseManager

app = FastAPI(title="Freelancer Financial Statement API", version="1.0.0")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# Initialize database manager
db_manager = FinancialDatabaseManager("financial_statement.db")

# Pydantic models for request/response
class LoginRequest(BaseModel):
    freelancer_id: int
    pin: str

class LoginResponse(BaseModel):
    freelancer_id: int
    freelancer_name: str
    is_superuser: bool

class FreelancerCreate(BaseModel):
    name: str
    email: str
    phone: str
    bank_name: str
    bank_account: str
    epf_percentage: float
    pin: str
    is_superuser: bool = False

class FreelancerResponse(BaseModel):
    id: int
    name: str
    email: str
    phone: str
    bank_name: str
    bank_account: str
    epf_percentage: float
    created_at: str

class ProjectCreate(BaseModel):
    freelancer_id: int
    project_name: str
    project_amount: float
    project_date: str
    notes: Optional[str] = ""

class ProjectResponse(BaseModel):
    id: int
    freelancer_id: int
    project_name: str
    project_amount: float
    epf_deduction: float
    project_date: str
    status: str
    created_at: str

class DeductionCreate(BaseModel):
    freelancer_id: int
    year: int
    month: int
    other_deduction: float = 0.0
    notes: Optional[str] = ""

class MonthlyDeductionResponse(BaseModel):
    id: int
    freelancer_id: int
    year: int
    month: int
    socso_deduction: float
    pcb_deduction: float
    other_deduction: float
    created_at: str

class FinancialSummaryResponse(BaseModel):
    freelancer: Optional[tuple]
    total_income: float
    epf_deduction: float
    socso_deduction: float
    pcb_deduction: float
    other_deduction: float
    total_deductions: float
    net_amount: float


# ==================== LOGIN ENDPOINT ====================

@app.post("/login/")
def login(request: LoginRequest):
    """Authenticate freelancer with ID and PIN"""
    result = db_manager.verify_login(request.freelancer_id, request.pin)
    
    if not result:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid Freelancer ID or PIN"
        )
    
    return LoginResponse(
        freelancer_id=result["freelancer_id"],
        freelancer_name=result["freelancer_name"],
        is_superuser=result["is_superuser"]
    )

# ==================== FREELANCER ENDPOINTS ====================

@app.post("/freelancers/")
def create_freelancer(freelancer: FreelancerCreate):
    """Create a new freelancer"""
    freelancer_id = db_manager.add_freelancer(
        name=freelancer.name,
        email=freelancer.email,
        phone=freelancer.phone,
        bank_name=freelancer.bank_name,
        bank_account=freelancer.bank_account,
        epf_percentage=freelancer.epf_percentage,
        pin=freelancer.pin,
        is_superuser=freelancer.is_superuser
    )
    
    if not freelancer_id:
        raise HTTPException(status_code=400, detail="Failed to add freelancer")
    
    return {"freelancer_id": freelancer_id, "message": "Freelancer created successfully"}

@app.get("/freelancers/")
def get_all_freelancers():
    """Get all freelancers"""
    freelancers = db_manager.get_all_freelancers()
    return freelancers

@app.get("/freelancers/{freelancer_id}")
def get_freelancer(freelancer_id: int):
    """Get specific freelancer"""
    freelancer = db_manager.get_freelancer(freelancer_id)
    
    if not freelancer:
        raise HTTPException(status_code=404, detail="Freelancer not found")
    
    return freelancer

# ==================== PROJECT ENDPOINTS ====================

@app.post("/projects/")
def create_project(project: ProjectCreate):
    """Create a new project"""
    project_id = db_manager.add_project(
        freelancer_id=project.freelancer_id,
        project_name=project.project_name,
        project_amount=project.project_amount,
        project_date=project.project_date,
        notes=project.notes
    )
    
    if not project_id:
        raise HTTPException(status_code=400, detail="Failed to add project")
    
    # Calculate EPF for response
    freelancer = db_manager.get_freelancer(project.freelancer_id)
    epf_deduction = project.project_amount * (freelancer["epf_percentage"] / 100) if freelancer else 0
    
    return {
        "project_id": project_id,
        "epf_deduction": epf_deduction,
        "message": "Project created successfully"
    }

@app.get("/freelancers/{freelancer_id}/projects/")
def get_freelancer_projects(freelancer_id: int):
    """Get all projects for a freelancer"""
    projects = db_manager.get_freelancer_projects(freelancer_id)
    
    return projects

# ==================== DEDUCTION ENDPOINTS ====================

@app.post("/deductions/")
def create_deduction(deduction: DeductionCreate):
    """Create monthly deductions"""
    deduction_id = db_manager.add_monthly_deduction(
        freelancer_id=deduction.freelancer_id,
        year=deduction.year,
        month=deduction.month,
        other=deduction.other_deduction,
        notes=deduction.notes
    )
    
    if not deduction_id:
        raise HTTPException(status_code=400, detail="Failed to add deductions")
    
    # Get calculated deductions
    summary = db_manager.get_freelancer_financial_summary(deduction.freelancer_id, deduction.year, deduction.month)
    
    return {
        "deduction_id": deduction_id,
        "socso_deduction": round(summary["socso_deduction"], 2) if summary else 0,
        "pcb_deduction": round(summary["pcb_deduction"], 2) if summary else 0,
        "message": "Deductions registered successfully"
    }

# ==================== FINANCIAL SUMMARY ENDPOINTS ====================

@app.get("/freelancers/{freelancer_id}/summary/{year}/{month}")
def get_monthly_summary(freelancer_id: int, year: int, month: int):
    """Get financial summary for a specific month"""
    summary = db_manager.get_freelancer_financial_summary(freelancer_id, year, month)
    
    if not summary:
        raise HTTPException(status_code=404, detail="No data found")
    
    return {
        "freelancer_name": summary['freelancer_name'],
        "total_income": round(summary['total_income'], 2),
        "epf_deduction": round(summary['epf_deduction'], 2),
        "socso_deduction": round(summary['socso_deduction'], 2),
        "pcb_deduction": round(summary['pcb_deduction'], 2),
        "other_deduction": round(summary['other_deduction'], 2),
        "total_deductions": round(summary['total_deductions'], 2),
        "net_amount": round(summary['net_amount'], 2)
    }

@app.get("/freelancers/{freelancer_id}/summary-range")
def get_date_range_summary(freelancer_id: int, start_date: str, end_date: str):
    """Get financial summary for a date range"""
    summary = db_manager.get_freelancer_financial_summary_range(freelancer_id, start_date, end_date)
    
    if not summary:
        raise HTTPException(status_code=404, detail="No data found")
    
    return {
        "freelancer_name": summary['freelancer'][0] if summary['freelancer'] else "N/A",
        "start_date": summary['start_date'],
        "end_date": summary['end_date'],
        "total_income": round(summary['total_income'], 2),
        "epf_deduction": round(summary['epf_deduction'], 2),
        "socso_deduction": round(summary['socso_deduction'], 2),
        "pcb_deduction": round(summary['pcb_deduction'], 2),
        "other_deduction": round(summary['other_deduction'], 2),
        "total_deductions": round(summary['total_deductions'], 2),
        "net_amount": round(summary['net_amount'], 2)
    }

# ==================== REPORT GENERATION ENDPOINTS ====================

@app.post("/generate-report/")
def generate_report(
    freelancer_id: int,
    report_type: str,  # "month", "year", "date_range"
    year: int = None,
    month: int = None,
    start_date: str = None,
    end_date: str = None,
    file_format: str = "csv",  # "csv" or "excel"
    include_fields: list = None  # fields to include in report
):
    """Generate financial statement report for download"""
    
    # Get freelancer info
    freelancer = db_manager.get_freelancer(freelancer_id)
    if not freelancer:
        raise HTTPException(status_code=404, detail="Freelancer not found")
    
    # Collect data based on report type
    if report_type == "month":
        summary = db_manager.get_freelancer_financial_summary(freelancer_id, year, month)
        if not summary:
            raise HTTPException(status_code=404, detail="No data found for this month")
        report_data = [summary]
        period_name = f"{month:02d}-{year}"
        
    elif report_type == "year":
        report_data = []
        for m in range(1, 13):
            summary = db_manager.get_freelancer_financial_summary(freelancer_id, year, m)
            if summary:
                report_data.append(summary)
        if not report_data:
            raise HTTPException(status_code=404, detail="No data found for this year")
        period_name = f"{year}"
        
    elif report_type == "date_range":
        summary = db_manager.get_freelancer_financial_summary_range(freelancer_id, start_date, end_date)
        if not summary:
            raise HTTPException(status_code=404, detail="No data found for this date range")
        report_data = [summary]
        period_name = f"{start_date}_to_{end_date}"
    else:
        raise HTTPException(status_code=400, detail="Invalid report type")
    
    # Default fields if none specified
    if not include_fields:
        include_fields = [
            "freelancer_name", "total_income", "epf_deduction",
            "socso_deduction", "pcb_deduction", "other_deduction",
            "total_deductions", "net_amount"
        ]
    
    # Generate file
    if file_format == "csv":
        output = generate_csv_report(report_data, include_fields, freelancer["name"], period_name)
        return StreamingResponse(
            iter([output]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=financial_statement_{period_name}.csv"}
        )
    
    elif file_format == "excel":
        if not HAS_OPENPYXL:
            raise HTTPException(status_code=400, detail="Excel export not available. Install openpyxl.")
        output = generate_excel_report(report_data, include_fields, freelancer["name"], period_name)
        return StreamingResponse(
            io.BytesIO(output),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=financial_statement_{period_name}.xlsx"}
        )
    
    elif file_format == "pdf":
        output = generate_pdf_report(report_data, include_fields, freelancer["name"], period_name)
        return StreamingResponse(
            io.BytesIO(output),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=financial_statement_{period_name}.pdf"}
        )
    
    else:
        raise HTTPException(status_code=400, detail="Invalid file format")

def generate_csv_report(data, fields, freelancer_name, period):
    """Generate CSV report"""
    output = io.StringIO()
    
    # Header
    output.write(f"Freelancer Financial Statement\n")
    output.write(f"Freelancer: {freelancer_name}\n")
    output.write(f"Period: {period}\n")
    output.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
    
    # Data
    writer = csv.DictWriter(output, fieldnames=fields)
    writer.writeheader()
    
    for row in data:
        filtered_row = {k: v for k, v in row.items() if k in fields}
        writer.writerow(filtered_row)
    
    return output.getvalue()

def generate_excel_report(data, fields, freelancer_name, period):
    """Generate Excel report"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Financial Statement"
    
    # Header style
    header_font = Font(bold=True, size=14)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF", size=14)
    
    # Title
    ws['A1'] = "Freelancer Financial Statement"
    ws['A1'].font = header_font_white
    ws['A1'].fill = header_fill
    ws.merge_cells('A1:H1')
    
    ws['A2'] = f"Freelancer: {freelancer_name}"
    ws['A3'] = f"Period: {period}"
    ws['A4'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    
    # Column headers
    col_num = 1
    header_fill_light = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font_bold = Font(bold=True)
    
    for field in fields:
        cell = ws.cell(row=6, column=col_num)
        cell.value = field.replace('_', ' ').title()
        cell.fill = header_fill_light
        cell.font = header_font_bold
        cell.alignment = Alignment(horizontal="center", vertical="center")
        col_num += 1
    
    # Data rows
    row_num = 7
    for row_data in data:
        col_num = 1
        for field in fields:
            cell = ws.cell(row=row_num, column=col_num)
            value = row_data.get(field, "")
            
            # Format currency
            if field in ['total_income', 'epf_deduction', 'socso_deduction', 'pcb_deduction', 'other_deduction', 'total_deductions', 'net_amount']:
                cell.value = float(value) if value else 0
                cell.number_format = '0.00'
            else:
                cell.value = value
            
            cell.alignment = Alignment(horizontal="right" if isinstance(value, (int, float)) else "left")
            col_num += 1
        row_num += 1
    
    # Auto-adjust columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def generate_pdf_report(data, fields, freelancer_name, period):
    """Generate PDF report"""
    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output, 
        pagesize=landscape(A4),
        topMargin=0.4*inch, 
        bottomMargin=0.4*inch, 
        leftMargin=0.25*inch, 
        rightMargin=0.25*inch
    )
    
    # Container for PDF elements
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=14,
        textColor=colors.HexColor('#1F4E78'),
        spaceAfter=4,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.black,
        spaceAfter=3
    )
    
    # Title
    title = Paragraph("Freelancer Financial Statement", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.15*inch))
    
    # Header info
    header_info = f"<b>Freelancer:</b> {freelancer_name}<br/>" \
                 f"<b>Period:</b> {period}<br/>" \
                 f"<b>Generated:</b> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    elements.append(Paragraph(header_info, heading_style))
    elements.append(Spacer(1, 0.1*inch))
    
    # Prepare table data
    table_data = [[]]
    
    # Add column headers
    for field in fields:
        header_text = field.replace('_', ' ').title()
        table_data[0].append(header_text)
    
    # Add data rows
    for row in data:
        row_data = []
        for field in fields:
            value = row.get(field, "")
            
            # Format currency
            if field in ['total_income', 'epf_deduction', 'socso_deduction', 'pcb_deduction', 'other_deduction', 'total_deductions', 'net_amount']:
                if value:
                    row_data.append(f"RM {float(value):,.2f}")
                else:
                    row_data.append("RM 0.00")
            else:
                row_data.append(str(value))
        
        table_data.append(row_data)

    # Calculate ultra-compact column widths
    # A4 available width with 0.25" margins = 7.77 inches
    available_width = 7.77 * inch
    num_cols = len(fields)
    
    # Use equal width for all columns
    col_width = available_width / num_cols
    col_widths = [col_width] * num_cols
    
    # Create table
    table = Table(table_data, colWidths=[1.2*inch] * len(fields))
    
    # Style table
    table.setStyle(TableStyle([
        # Header row style
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 7),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 4),
        ('TOPPADDING', (0, 0), (-1, 0), 4),
        
        # Data rows style
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 6.5),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 1), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
        
        # Align currency columns to right
        ('ALIGN', (0, 1), (-1, -1), 'RIGHT'),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 0.15*inch))
    
    # Footer
    footer_text = "This document is confidential and for authorized use only."
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=7,
        textColor=colors.grey,
        alignment=TA_CENTER
    )
    elements.append(Paragraph(footer_text, footer_style))
    
    # Build PDF
    doc.build(elements)
    output.seek(0)
    return output.getvalue()

# ==================== DELETE ENDPOINTS ====================

@app.delete("/freelancers/{freelancer_id}", response_model=Dict)
async def delete_freelancer(freelancer_id: int):
    """Delete a freelancer and all related data"""
    try:
        # Check if freelancer exists
        with sqlite3.connect(db_manager.db_name) as conn:
            cursor = conn.cursor()
            cursor.execute(
                "SELECT id FROM freelancers WHERE id = ?",
                (freelancer_id,)
            )
            if not cursor.fetchone():
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="Freelancer not found"
                )

        if db_manager.delete_freelancer(freelancer_id):
            return {
                "message": "Freelancer deleted successfully",
                "freelancer_id": freelancer_id
            }
        else:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Failed to delete freelancer"
            )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Internal server error: {str(e)}"
        )
    
# ==================== STATUS ENDPOINT ====================

@app.get("/")
def api_status():
    """Check API status"""
    return {
        "status": "running",
        "version": "1.0.0",
        "message": "Freelancer Financial Statement API"
    }

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)